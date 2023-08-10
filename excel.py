from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

#Datalar. İstediğiniz Gibi Oluşturup Değer Verebilirsiniz.
data = {
 "Bilgi": {
  "Bilgi1": 14,
  "Bilgi2": 13,
  "Bilgi3": 13,
  "Bilgi4": 31
 },
 "Bilgi2": {
  "Bilgi1": 51,
  "Bilgi2": 145,
  "Bilgi3": 12,
  "Bilgi4": 12
 },
  "Bilgi3":{
  "Bilgi1": 1,
  "Bilgi2": 12,
  "Bilgi3": 2,
  "Bilgi4": 23
 },
}

w1 = Workbook()
w2 = w1.active
w2.title = "Grades" #başlık
headings = ['Name'] + list(data['Bilgi'].keys())
w2.append(headings)

for person in data:
 grades = list(data[person].values())
 w2.append([person] + grades)

for col in range(2, len(data['Bilgi']) + 2):
 char = get_column_letter(col)
 w2[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
 w2[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

w1.save("Yeniexcel.xlsx") #dosya oluşum yeri başlığı kendiniz verin.

